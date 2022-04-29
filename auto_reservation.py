from selenium import webdriver
import logging
import logging.handlers
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.action_chains import ActionChains
import datetime
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import pyautogui as pag
import openpyxl as pyxl
import pandas as pd

now = datetime.datetime.now()

# Logの設定
logger=logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
h1=logging.handlers.RotatingFileHandler('Log.txt', maxBytes=100000, backupCount=10)
h1.setLevel(logging.DEBUG)
formatter=logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
h1.setFormatter(formatter)
logger.addHandler(h1)


# 日時選択・空き状況確認関数
def check_open(driver, day, time):

    logger.debug('-- day : {} --'.format(day))
    logger.debug('-- time : {} --'.format(time))
    
    # time = time.split(":")
    # if int(time[1]) != 0:
    #     time = time[0] + ":" + format(str(int(time[1]) - 30), "0>2")
    # else:
    #     time = format(str(int(time[0]) -1), "0>2") + ":" + format(str(int(time[1]) + 30), "0>2")
    day = "'" + day + "'"
    time = "'" + time + "-" + "'"
    # time = "'" + time + "'"
    # logger.debug('-- day 2: {} --'.format(day))
    # logger.debug('-- time 2: {} --'.format(time))
    # 予約受付開始前の状態を確認
    # not_open_path = "//td[@data-name_message='お電話にてお問い合わせください']"
    # NotOpenElement = driver.find_elements_by_xpath(not_open_path)
    # if len(NotOpenElement) < 0:
    #     return None, not_open_path

    # else:
    for mark in ["◎", "○", "△"]:
        mark = "'" + mark + "'"
        logger.debug("-- mark : {} --".format(mark))
        # open_path = "//tr[th[@data-sys_time = {0}]]/td[input[@value = {1}]]/span[contains(text(), '◎')]".format(time, day)
        open_path = "//tr[th[@data-sys_time={0}]]/td[input[@value = {1}]]/span[contains(text(), {2})]".format(time, day, mark)
        logger.debug('-- open path = {} --'.format(open_path))
        OpenElement = driver.find_elements_by_xpath(open_path)
        if len(OpenElement) > 0:
            logger.debug('-- 予約を進めます --')
            return OpenElement, open_path
        else:
            if mark == "'△'":
                logger.debug('-- 予約できません。 --')
                print("予約できません。")
                OpenElement = None
                return OpenElement, open_path
            else:
                continue
    # open_path = open_path.replace("/span[contains(text(), '◎')]", "")
    # open_path = "//tr[th[contains(text(), {0})]]/td/input[@value = {1}]".format(time, day)

def select_object(driver, name, age):

    selected_time = driver.find_elements_by_xpath("//span[@class='disp-datetime']/div[@class='disp-time']")[0].text
    logger.debug('-- selected time -- {}'.format(selected_time))
    if selected_time != times[int(i)]: # 希望時間と選択時間が一致しない場合
        if datetime.datetime.strptime(selected_time, "%H:%M") > datetime.datetime.strptime(str(times[int(i)]), "%H:%M"):
            # 選択時間の方が後の場合
            driver.find_elements_by_xpath("//a[@class='arrow-prev-btn']")[0].click()
        else:
            driver.find_elements_by_xpath("//a[@class='arrow-next-btn']")[0].click()
        selected_time_2 = driver.find_elements_by_xpath("//span[@class='disp-datetime']/div[@class='disp-time']")[0].text
        logger.debug('-- selected time_2 -- {}'.format(selected_time_2))

    # wait.until(EC.visibility_of_all_elements_located)

    if name != "":
        logger.debug('-- 指名あり --')
        name = "'" + name + "(" + str(age) + "歳" + ")" + "'"
        logger.debug('-- name : {} --'.format(name))
        element_object_id = driver.find_elements_by_xpath("//a[strong[contains(text(), {0})]]".format(name))
        if len(element_object_id) > 0:
            logger.debug('-- 対象が見つかりました。 --')
            object_id = str(element_object_id[0].get_attribute("href").split('/')[-1].split('-')[-1])
            logger.debug('-- object_id : {} --'.format(object_id))
            element_select_object = driver.find_elements_by_xpath("//a[input[@value = {}]]".format(object_id))
            logger.debug("-- element_select_object : {} --".format(element_select_object))
            return element_select_object
        
        else:
            logger.debug('-- 対象が見つかりませんでした。 --')
            return None
    else:
        logger.debug('-- 指名なし --')
        element_object_id = driver.find_elements_by_xpath("//a[contains(text(), ' 指名なし（フリー予約）                                ')]")
        logger.debug('-- element_object_id : {} --'.format(element_object_id))
        return element_object_id
    
def create_time_range(start, end):
    list = []
    list.append(start)
    start = datetime.datetime.strptime(start, "%H:%M")
    end = datetime.datetime.strptime(end, "%H:%M")
    if start == end or end == None:
        list = [format(str(start.hour), "0>2") + ":" + format(str(start.minute), "0>2")]
        logger.debug('-- time-list -- {}'.format(list))
        return list

    minutes = int((end - start) / datetime.timedelta(minutes=30))
    for minute in range(0, minutes):
        start = start + datetime.timedelta(minutes=30)
        list.append(format(str(start.hour), "0>2") + ":" + format(str(start.minute), "0>2"))
        if start == end:
            logger.debug("-- time-list -- {}".format(list))
            return list


def select_course(driver, hour, find_object):

    if find_object != None:

        find_object[0].click()

        # wait.until(EC.visibility_of_all_elements_located)
        sleep(2)
        
        hour = "'" + str(hour) + "分" + "'"
        hour_path = "//div[span[strong[contains(text(), {0})]]]/following-sibling::div/span[input[@value='選択する']]".format(hour)
        # hour_path = "//div[span[strong[contains(text(), {0})]]]/following-sibling::div/span[@class='select_btn']".format(hour)
        logger.debug('-- hour -- {}'.format(hour))
        logger.debug('-- hour path -- {}'.format(hour_path))
        selected_hour = driver.find_elements_by_xpath(hour_path)
        if len(selected_hour) > 0:

            return selected_hour

        else:
            logger.debug('-- course not found --')
            return None

def login(driver, id, password):

    if len(driver.find_elements_by_xpath("//ul[@id='login_item']")) > 0:

        # ID/メールアドレス
        driver.find_elements_by_xpath("//ul[@id='login_item']/li/input[@id='user']")[0].send_keys(id)

        # Password
        driver.find_elements_by_xpath("//ul[@id='login_item']/li/input[@id='pass']")[0].send_keys(password)

        #ログインボタン
        driver.find_elements_by_xpath("//div[@class='loginButton']")[0].click()


def accept_terms(driver):

    if len(driver.find_elements_by_xpath("//h2[contains(text(), '予約システムの流れとご利用規約')]")) > 0:


        # path = driver.find_elements_by_xpath("//div[@class='term-input']/button[contains(text(), '次へ')]")
        path = driver.find_elements_by_xpath("//button[contains(text(), '次へ')]")
        # actions.move_to_element(path[0])
        sleep(2)
        # actions.double_click(path[0]).perform()
        logger.debug('-- 利用規約に同意します。 --')
        path[0].click()
        path[0].click()

def complete_reservation(driver):

    logger.debug('-- 予約を完了します。 --')
    # driver.find_elements_by_xpath("//a[@id='reservation_confirm']")[0].click() # 最終動作確認までコメントアウトしておく
    pag.alert(driver.find_elements_by_xpath("//a[@id='reservation_confirm']")[0].text)

# def open_hour(name, start="6", end="24"):
#     logger.debug("-- 営業時間設定 --")
#     print(start)
#     print(end)
#     minutes = end - start
#     minutes = minutes / 30
#     for minute in range(1, minutes):

#     print(minutes)

# 事前情報の設定
wb = pyxl.load_workbook("予約.xlsx")
ws = wb['Sheet1']
df=pd.read_excel("予約.xlsx", sheet_name='Sheet1')
objects = []
for row in range(2, len(df)):
    row = str(row)
    name = ws["C{}".format(row)].value
    age = ws["D{}".format(row)].value
    url = ws["B{}".format(row)].value
    objects.append({"name":name, "age":int(age), "url":str(url)})

# objects = [{"name":"さく", "age":20}, {"name":"ちなつ", "age":19}, {"name":"みつば", "age":21},  {"name":"えの", "age":20}, {"name":"ひめか", "age":19}, {"name":"こむぎ", "age":20}, {"name":"さくら", "age":19}]
logger.debug('-- objects -- {}'.format(objects))
day = "2022-04-29"
time = {"start":"15:30", "end":"19:00"}
times = create_time_range(start=time["start"], end=time["end"])
hour = 50
id = "poko576@yahoo.co.jp"
password = "wa42phzz"

booked = ""

def start_driver(url):
    logger.debug('-- WebDriverの設定 --')
    # try:
    # DriverManager : https://jpdebug.com/p/2615980
    # Chromeの場合
    options = Options()
    options.add_argument('--headless')                         
    options.add_argument('--disable-gpu')                      
    options.add_argument('--disable-extensions')               
    options.add_argument('--proxy-server="direct://"')         
    options.add_argument('--proxy-bypass-list=*')              
    options.add_argument('--blink-settings=imagesEnabled=false')
    options.add_argument('--lang=ja')                          
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument("--log-level=3")
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('--start-maximized')
    options.page_load_strategy = 'eager'

    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get(url)
    # driver.get('https://www.cityheaven.net/kanagawa/A1403/A140301/afterschool/A6ShopReservation/') # お店予約ページ
    # driver.get('https://www.cityheaven.net/kanagawa/A1403/A140301/afterschool/') # お店トップページ

    # 参考サイト: https://tetoblog.org/2021/05/disney-scraping/
    # 参考サイト２: https://qiita.com/memakura/items/20a02161fa7e18d8a693

    # 待機時間の設定
    logger.debug('-- 待機時間の設定 --')
    wait = WebDriverWait(driver, 20) # 最大待機時間 10 秒
    # wait.until(EC.visibility_of_all_elements_located)
    # sleep(10)

    # ページ要素の取得
    logger.debug('-- ページ要素の取得 --')
    # button = driver.find_element_by_id('reserve_btn')
    # button.click()
    sleep(2)
    # wait.until(EC.visibility_of_all_elements_located)

    # サイトにiframeが使われているときは、iframeに出入りする必要がある。
    # https://qiita.com/hondy12345/items/c7359c300c73afd11d76
    # ここでiframeに入る
    logger.debug('-- enter iframe --')
    iframe = driver.find_element_by_id('pcreserveiframe')
    driver.switch_to.frame(iframe)
    return driver
    # iframeから抜けるコード
    # driver.switch_to.default_content()

driver = start_driver(url=objects[0]["url"])
n = 0

try:
    while booked != "Done":
        for object in objects:
            n += 1
            logger.debug("-- object -- {}".format(object))

            if booked == "Done":
                break
            
            for i in range(len(times)):

                if booked == "Done":
                    break


                # check_open関数で予約がopenになっているか確認
                # actions = ActionChains(driver)
                logger.debug("-- check_open関数 --")
                element, xpath = check_open(driver=driver, day=day, time=times[int(i)])
                print("elements : ", element, " xpath : ", xpath)
                logger.debug('-- elements : ', element, " xpath : ", xpath, " --")
                
                if element == None:
                    continue
                elif element != None:
                    result = driver.find_element_by_xpath(xpath)
                    # actions.move_to_element(result)
                    sleep(1)
                    result.click()
                    sleep(3)
                    # wait.until(EC.visibility_of_all_elements_located)


                    # 対象を選択する
                    logger.debug('-- select_object関数 --')
                    find_object = select_object(driver=driver, name=object["name"], age=object["age"])

                    # 対象が存在する場合
                    # コースの選択
                    logger.debug('-- select_course関数 --')
                    selected_course = select_course(driver=driver, hour=hour, find_object=find_object)

                    if selected_course == None:
                        driver.close()
                        driver = start_driver(url=objects[n]["url"])
                        continue
                    elif selected_course != None:

                        # actions.move_to_element(selected_course[0])
                        selected_course[0].click()
                        # wait.until(EC.visibility_of_all_elements_located)
                        sleep(3)
                        logger.debug('-- Course Select Completed --')

                        # 確定して次へ
                        logger.debug('-- 確定して次へ --')
                        next = driver.find_elements_by_xpath("//div[@class='booking-footer']/button[@title='確定して次へ']")
                        if len(next) > 0:
                            next[0].click()
                            # wait.until(EC.visibility_of_all_elements_located)
                            sleep(2)

                        # ログインが要求されれば、ログインを行う
                        logger.debug('-- ログイン情報入力 --')
                        login(driver=driver, id=id, password=password)
                        # wait.until(EC.visibility_of_all_elements_located)
                        sleep(2)

                        # 利用規約への同意    
                        logger.debug('-- 利用規約の同意 --')
                        accept_terms(driver=driver)
                        # wait.until(EC.visibility_of_all_elements_located)
                        sleep(2)

                        # 連絡先を確定して次へ
                        logger.debug('-- 連絡先を確定 --')
                        contact = driver.find_elements_by_id('next')[0]
                        # actions.move_to_element(contact)
                        contact.click()
                        contact.click()
                        # wait.until(EC.visibility_of_all_elements_located)
                        sleep(2)
                    
                        # 予約完了
                        logger.debug('-- 予約完了 --')
                        complete_reservation(driver=driver)
                        sleep(1)

                        booked = "Done"

                        driver.close()

except:
    driver.close()
    logger.debug('-- WebDriverの設定 --')
    # try:
    # DriverManager : https://jpdebug.com/p/2615980
    # Chromeの場合
    options = Options()
    options.add_argument('--headless')                         
    options.add_argument('--disable-gpu')                      
    options.add_argument('--disable-extensions')               
    options.add_argument('--proxy-server="direct://"')         
    options.add_argument('--proxy-bypass-list=*')              
    options.add_argument('--blink-settings=imagesEnabled=false')
    options.add_argument('--lang=ja')                          
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument("--log-level=3")
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('--start-maximized')
    options.page_load_strategy = 'eager'

    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get('https://www.cityheaven.net/kanagawa/A1403/A140301/afterschool/A6ShopReservation/') # お店予約ページ
    # driver.get('https://www.cityheaven.net/kanagawa/A1403/A140301/afterschool/') # お店トップページ

    # 参考サイト: https://tetoblog.org/2021/05/disney-scraping/
    # 参考サイト２: https://qiita.com/memakura/items/20a02161fa7e18d8a693

    # 待機時間の設定
    logger.debug('-- 待機時間の設定 --')
    wait = WebDriverWait(driver, 20) # 最大待機時間 10 秒
    # wait.until(EC.visibility_of_all_elements_located)
    # sleep(10)

    # ページ要素の取得
    logger.debug('-- ページ要素の取得 --')
    # button = driver.find_element_by_id('reserve_btn')
    # button.click()
    sleep(2)
    # wait.until(EC.visibility_of_all_elements_located)

    # サイトにiframeが使われているときは、iframeに出入りする必要がある。
    # https://qiita.com/hondy12345/items/c7359c300c73afd11d76
    # ここでiframeに入る
    logger.debug('-- enter iframe --')
    iframe = driver.find_element_by_id('pcreserveiframe')
    driver.switch_to.frame(iframe)
    # iframeから抜けるコード
    # driver.switch_to.default_content()

                        
                    
        

    


# pag.alert('ALL COMPLETED')


# BeautifulSoupでtable要素を取得するパターン（今回は使わない）
# html = driver.page_source
# soup = BS(html, "html.parser")

# # カレンダーから予約のページ
# table = soup.findAll("table", {"class":"cth table table-bordered table-responsive concat_table table-fixed"})
# table = table[0]
# table = table.findAll("tr")

# for i, tr in enumerate(table):
#     print(tr.get_text())
#     tmp = tr.findAll("td")
#     for k, td in enumerate(tmp):
#         print(td.get_text())



# except Exception:
#     print(Exception)
#     driver.close()


